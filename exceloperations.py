from openpyxl import Workbook
wb=Workbook()
sheet=wb.active

lang=['kannada','telugu','tamil']
state=['karnataka','andra','tamilnadu']
code=['KA','TS','TN']
cap=['blore','hyd','chennai']

sheet.cell(1,1).value='lang'
sheet.cell(1,2).value='state'
sheet.cell(1,3).value='code'
sheet.cell(1,4).value='cap'

for i in range(2,5):
    sheet.cell(i, 1).value = lang[i-2]
    sheet.cell(i, 2).value = state[i - 2]
    sheet.cell(i, 3).value = code[i - 2]
    sheet.cell(i, 4).value = cap[i - 2]
c=input("enter the code")

for i in range(2,5):
    if c==sheet.cell(i,3).value:
        print (sheet.cell(i,1).value ,'and',sheet.cell(i,4).value)




